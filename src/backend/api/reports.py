"""
Module 05 - Reporting & Analytics
Endpoints for generating and exporting reports in PDF and Excel formats
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime
from io import BytesIO

from auth.auth import get_current_user, User
from data_store import rh_employees, rh_contracts, collab_leave_requests

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

router = APIRouter()

# ============================================
# REPORT DATA ENDPOINTS
# ============================================

@router.get("/etat-personnel")
def get_etat_personnel(
    department_id: Optional[int] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get employee status report data"""
    employees = rh_employees
    
    if department_id:
        employees = [e for e in employees if e.department_id == department_id]
    if status:
        employees = [e for e in employees if e.status == status]
    
    return {
        "generated_at": datetime.now().isoformat(),
        "total_employees": len(employees),
        "employees": [
            {
                "matricule": e.matricule,
                "nom": f"{e.first_name} {e.last_name}",
                "poste": e.position,
                "departement": f"Département {e.department_id}" if e.department_id else "Non assigné",
                "type_contrat": e.contract_type,
                "date_embauche": e.hire_date,
                "statut": e.status,
                "email": e.professional_email,
                "telephone": e.professional_phone
            }
            for e in employees
        ]
    }

@router.get("/conges")
def get_report_conges(
    year: int = Query(default=2024, ge=2000, le=2100),
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Get leave report data"""
    requests = collab_leave_requests
    
    return {
        "generated_at": datetime.now().isoformat(),
        "period": f"{month}/{year}" if month else str(year),
        "total_requests": len(requests),
        "approved": len([r for r in requests if r.status == 'approved']),
        "pending": len([r for r in requests if r.status == 'pending']),
        "requests": [
            {
                "id": r.id,
                "type": r.type,
                "date_debut": r.start,
                "date_fin": r.end,
                "jours": r.days,
                "statut": r.status
            }
            for r in requests
        ]
    }

@router.get("/contrats")
def get_report_contrats(
    contract_type: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get contracts report data"""
    contracts = rh_contracts
    if contract_type:
        contracts = [c for c in contracts if c.type == contract_type]

    return {
        "generated_at": datetime.now().isoformat(),
        "total_contracts": len(contracts),
        "active": len([c for c in contracts if c.status == 'active']),
        "expiring_soon": len([c for c in contracts if c.alert]),
        "contracts": [
            {
                "id": c.id,
                "employe": c.employee,
                "type": c.type,
                "debut": c.start,
                "fin": c.end,
                "statut": c.status,
                "alerte": c.alert
            }
            for c in contracts
        ]
    }

# ============================================
# PDF EXPORT ENDPOINTS
# ============================================

def generate_pdf_report(title: str, headers: List[str], data: List[List], subtitle: str = ""):
    """Helper function to generate PDF report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e3a5f'),
        spaceAfter=20
    )
    elements.append(Paragraph(title, title_style))
    
    # Subtitle
    if subtitle:
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=10
        )
        elements.append(Paragraph(f"Généré le: {subtitle}", subtitle_style))
    
    elements.append(Spacer(1, 20))
    
    # Table
    table_data = [headers] + data
    table = Table(table_data, repeatRows=1)
    
    # Table style
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')])
    ]))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1  # Center
    )
    elements.append(Paragraph("© ICES RH - Document confidentiel", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

@router.get("/export/etat-personnel/pdf")
def export_etat_personnel_pdf(
    department_id: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Export employee status report as PDF"""
    employees = rh_employees
    if department_id:
        employees = [e for e in employees if e.department_id == department_id]
    
    headers = ["Matricule", "Nom", "Poste", "Département", "Contrat", "Date embauche", "Statut"]
    data = [
        [
            e.matricule,
            f"{e.first_name} {e.last_name}",
            e.position,
            f"Dépt {e.department_id}" if e.department_id else "N/A",
            e.contract_type,
            e.hire_date,
            e.status
        ]
        for e in employees
    ]
    
    pdf_buffer = generate_pdf_report(
        "État du Personnel",
        headers,
        data,
        datetime.now().strftime("%d/%m/%Y %H:%M")
    )
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=etat_personnel.pdf"}
    )

@router.get("/export/conges/pdf")
def export_conges_pdf(
    year: int = Query(default=2024),
    current_user: User = Depends(get_current_user)
):
    """Export leave report as PDF"""
    requests = collab_leave_requests
    
    headers = ["ID", "Type", "Date début", "Date fin", "Jours", "Statut"]
    data = [
        [str(r.id), r.type, r.start, r.end, str(r.days), r.status]
        for r in requests
    ]
    
    pdf_buffer = generate_pdf_report(
        f"Rapport des Congés - {year}",
        headers,
        data,
        datetime.now().strftime("%d/%m/%Y %H:%M")
    )
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=rapport_conges_{year}.pdf"}
    )

@router.get("/export/contrats/pdf")
def export_contrats_pdf(
    current_user: User = Depends(get_current_user)
):
    """Export contracts report as PDF"""
    contracts = rh_contracts
    
    headers = ["ID", "Employé", "Type", "Début", "Fin", "Statut", "Alerte"]
    data = [
        [
            str(c.id),
            c.employee,
            c.type,
            c.start,
            c.end or "-",
            c.status,
            c.alert or "-"
        ]
        for c in contracts
    ]
    
    pdf_buffer = generate_pdf_report(
        "Rapport des Contrats",
        headers,
        data,
        datetime.now().strftime("%d/%m/%Y %H:%M")
    )
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=rapport_contrats.pdf"}
    )

# ============================================
# EXCEL EXPORT ENDPOINTS
# ============================================

def generate_excel_report(title: str, headers: List[str], data: List[List], sheet_name: str = "Données"):
    """Helper function to generate Excel report"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Title
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color="1E3A5F")
    ws.merge_cells(f'A1:{chr(64+len(headers))}1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Date
    ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(size=9, italic=True, color="666666")
    ws.merge_cells(f'A2:{chr(64+len(headers))}2')
    
    # Headers
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@router.get("/export/etat-personnel/excel")
def export_etat_personnel_excel(
    department_id: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Export employee status report as Excel"""
    employees = rh_employees
    if department_id:
        employees = [e for e in employees if e.department_id == department_id]
    
    headers = ["Matricule", "Nom", "Poste", "Département", "Contrat", "Date embauche", "Statut", "Email", "Téléphone"]
    data = [
        [
            e.matricule,
            f"{e.first_name} {e.last_name}",
            e.position,
            f"Département {e.department_id}" if e.department_id else "Non assigné",
            e.contract_type,
            e.hire_date,
            e.status,
            e.professional_email,
            e.professional_phone or "-"
        ]
        for e in employees
    ]
    
    excel_buffer = generate_excel_report("État du Personnel", headers, data, "Personnel")
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=etat_personnel.xlsx"}
    )

@router.get("/export/conges/excel")
def export_conges_excel(
    year: int = Query(default=2024),
    current_user: User = Depends(get_current_user)
):
    """Export leave report as Excel"""
    requests = collab_leave_requests
    
    headers = ["ID", "Type de congé", "Date début", "Date fin", "Nombre de jours", "Statut"]
    data = [
        [r.id, r.type, r.start, r.end, r.days, r.status]
        for r in requests
    ]
    
    excel_buffer = generate_excel_report(f"Rapport des Congés - {year}", headers, data, "Congés")
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rapport_conges_{year}.xlsx"}
    )

@router.get("/export/contrats/excel")
def export_contrats_excel(
    current_user: User = Depends(get_current_user)
):
    """Export contracts report as Excel"""
    contracts = rh_contracts
    
    headers = ["ID", "Employé", "Type de contrat", "Date début", "Date fin", "Statut", "Alerte"]
    data = [
        [c.id, c.employee, c.type, c.start, c.end or "-", c.status, c.alert or "-"]
        for c in contracts
    ]
    
    excel_buffer = generate_excel_report("Rapport des Contrats", headers, data, "Contrats")
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rapport_contrats.xlsx"}
    )

# ============================================
# KPIs DASHBOARD DATA
# ============================================

@router.get("/kpis/tableau-de-bord")
def get_kpis_dashboard(
    current_user: User = Depends(get_current_user)
):
    """Get KPIs for real-time dashboard"""
    employees = rh_employees
    contracts = rh_contracts
    leave_requests = collab_leave_requests

    total_employees = len(employees)
    active_employees = len([e for e in employees if e.status == 'active'])
    on_leave = len([e for e in employees if e.status == 'on_leave'])
    
    total_contracts = len(contracts)
    expiring_contracts = len([c for c in contracts if c.alert])
    
    pending_leaves = len([r for r in leave_requests if r.status == 'pending'])
    approved_leaves = len([r for r in leave_requests if r.status == 'approved'])
    
    return {
        "generated_at": datetime.now().isoformat(),
        "effectifs": {
            "total": total_employees,
            "actifs": active_employees,
            "en_conge": on_leave,
            "taux_presence": round((active_employees / total_employees * 100), 1) if total_employees > 0 else 0
        },
        "contrats": {
            "total": total_contracts,
            "en_alerte": expiring_contracts,
            "taux_renouvellement": round(((total_contracts - expiring_contracts) / total_contracts * 100), 1) if total_contracts > 0 else 0
        },
        "conges": {
            "demandes_en_attente": pending_leaves,
            "approuvees": approved_leaves,
            "total_demandes": len(leave_requests)
        },
        "tendances": {
            "evolution_effectif": "+5%",
            "turnover": "8%",
            "absenteisme": "3.2%"
        }
    }

@router.get("/productivite")
def get_report_productivity(
    team_id: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Module 08 - Performance & Productivity Data"""
    # En production, ces KPIs seraient calculés ou tirés d'une table dédiée
    kpis = []
    return {
        "generated_at": datetime.now().isoformat(),
        "kpis": kpis,
        "alerts": ["Retard critique sur Projet Migration AWS"] if not team_id else []
    }

@router.get("/sanctions")
def get_report_disciplinaire(current_user: User = Depends(get_current_user)):
    """Legal/Disciplinary report for RH/Direction only"""
    if current_user.role not in ["admin_rh", "resp_rh", "direction"]:
        raise HTTPException(status_code=403, detail="Accès réservé")
    
    return {
        "generated_at": datetime.now().isoformat(),
        "records": [
            {"employe": "Emma Richard", "type": "Avertissement", "motif": "Retards répétés", "date": "10/05/2024"}
        ]
    }

@router.get("/export/salaires/excel")
def export_salaires_excel(current_user: User = Depends(get_current_user)):
    """Module 05 - Rapport des salaires (Masse salariale)"""
    if current_user.role not in ["admin_rh", "resp_rh", "direction"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    employees = rh_employees
    headers = ["Matricule", "Employé", "Département", "Salaire Base (FCFA)", "Primes", "Brut", "Net"]
    data = [
        [e.matricule, f"{e.first_name} {e.last_name}", f"D{e.department_id}", e.base_salary, 0, e.base_salary, e.base_salary * 0.8]
        for e in employees
    ]
    
    excel_buffer = generate_excel_report("État de la Masse Salariale", headers, data, "Salaires")
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rapport_salaires.xlsx"}
    )

@router.get("/export/pointage/pdf")
def export_pointage_pdf(date_debut: str, date_fin: str, current_user: User = Depends(get_current_user)):
    """Module 05 - Rapport de présence et ponctualité"""
    logs = []
    
    headers = ["Date", "Employé", "Arrivée", "Départ", "Lieu", "Statut"]
    data = [
        [log.get('date', ''), f"Employé #{log.get('employee_id', '')}", log.get('clock_in', ''), log.get('clock_out', '') or "-", log.get('location', ''), "Présent"]
        for log in logs
    ]
    
    pdf_buffer = generate_pdf_report(
        f"Rapport de Présence du {date_debut} au {date_fin}",
        headers,
        data,
        datetime.now().strftime("%d/%m/%Y %H:%M")
    )
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=rapport_presence.pdf"}
    )

@router.get("/export/etat-cnss/excel")
def export_cnss_excel(current_user: User = Depends(get_current_user)):
    """Module 05 - Export pour déclaration CNSS (Bénin)"""
    if current_user.role not in ["admin_rh", "resp_rh"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
        
    employees = rh_employees
    headers = ["N° CNSS", "Matricule", "Nom & Prénom", "Salaire Brut", "Part Ouvrière (3.6%)", "Part Patronale (21.4%)"]
    data = [
        [
            e.get('num_cnss') or "N/A",
            e.matricule,
            f"{e.first_name} {e.last_name}",
            e.base_salary,
            e.base_salary * 0.036,
            e.base_salary * 0.214
        ]
        for e in employees
    ]
    
    excel_buffer = generate_excel_report("Déclaration Nominative CNSS", headers, data, "CNSS")
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=declaration_cnss.xlsx"}
    )
